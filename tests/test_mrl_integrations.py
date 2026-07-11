"""Sprint 15.5 · integraciones Excel MRL."""

from __future__ import annotations

import unittest
from datetime import date
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.inventario_comercial.exports import excel_compras_mrl, excel_ventas_mrl
from app.maintenance.mrl_exports import export_ordenes_excel


def empresa():
    return SimpleNamespace(id=1, razon_social="Empresa Demo", nit="900-1", zona_horaria="America/Bogota", moneda="COP")


class TestMRLIntegrations(unittest.TestCase):
    def assert_mrl_workbook(self, result, expected_header):
        content, filename = result
        self.assertTrue(filename.endswith(".xlsx"))
        ws = load_workbook(BytesIO(content)).active
        self.assertEqual(ws["A1"].value, "Empresa Demo")
        self.assertEqual(ws["A7"].value, expected_header)

    def test_orders_export(self):
        machine = SimpleNamespace(codigo="EQ-1", nombre="Bomba", area="A", ubicacion="U")
        order = SimpleNamespace(id=1, numero="OT-1", titulo="Revisar", machine=machine, area="", ubicacion="", tipo="correctivo", status="abierta", prioridad="media", fecha_programada=date(2026, 7, 11), fecha_inicio=None, fecha_cierre=None, tiempo_gastado_label="—", technician=None)
        self.assert_mrl_workbook(export_ordenes_excel(empresa(), [order]), "Número")

    def test_purchases_export(self):
        item = SimpleNamespace(id=1, numero="CO-1", fecha=date(2026, 7, 11), fecha_factura=None, fecha_vencimiento=None, proveedor=None, moneda_factura="COP", subtotal=100, monto_iva=19, total=119, monto_pagado=0, saldo_pendiente=119, estado_pago_label="Pendiente", lineas=[SimpleNamespace(cantidad=2)])
        self.assert_mrl_workbook(excel_compras_mrl(empresa(), [item]), "Número")

    def test_sales_export(self):
        item = SimpleNamespace(id=1, numero="VE-1", fecha=date(2026, 7, 11), cliente=None, es_credito=True, estado_cobro_label="Pendiente", moneda="COP", total=50, monto_cobrado=10, saldo_pendiente=40, fecha_vencimiento=None, lineas=[SimpleNamespace(cantidad=1)])
        self.assert_mrl_workbook(excel_ventas_mrl(empresa(), [item]), "Número")


if __name__ == "__main__":
    unittest.main()
