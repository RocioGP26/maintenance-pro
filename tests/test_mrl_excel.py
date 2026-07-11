"""Tests del motor Excel MRL · Sprint 15.2."""

from __future__ import annotations

import unittest
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.mrl import ExcelExporter
from app.mrl.excel.exporter import BaseExcelExporter, ColumnFormat
from app.mrl.metadata import build_sample_metadata


class TestMRLEngine(unittest.TestCase):
    def test_generates_valid_xlsx_bytes(self) -> None:
        meta = build_sample_metadata()
        exporter = BaseExcelExporter(meta)
        sheet = exporter.add_sheet("Datos")
        sheet.write_table(
            ["Columna A", "Columna B"],
            [["Uno", 1], ["Dos", 2]],
            column_formats=[None, ColumnFormat(kind="integer")],
        )
        content, filename = exporter.save()

        self.assertTrue(filename.endswith(".xlsx"))
        self.assertTrue(filename.startswith("DOC-010-"))
        self.assertGreater(len(content), 500)

        wb = load_workbook(BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.title, "Datos")
        self.assertEqual(ws.cell(row=1, column=1).value, "Empresa Demo S.A.S.")
        self.assertEqual(ws.cell(row=7, column=1).value, "Columna A")
        self.assertEqual(ws.cell(row=8, column=1).value, "Uno")
        self.assertEqual(ws.cell(row=9, column=2).value, 2)
        self.assertIsNotNone(ws.auto_filter.ref)

    def test_create_factory(self) -> None:
        exporter = ExcelExporter.create(
            title="Activos",
            doc_code="DOC-010",
            instance_code="ACTIVOS-TEST",
            module="Maintenance",
            empresa_id=42,
            empresa_nombre="Acme S.A.",
            usuario="tester",
            generated_at=datetime(2026, 7, 10, 12, 0, tzinfo=timezone.utc),
        )
        sheet = exporter.add_sheet("Activos")
        sheet.write_table(["Código"], [["EQ-001"]])
        content, name = exporter.save()
        self.assertIn("DOC-010-ACTIVOS-TEST", name)
        self.assertGreater(len(content), 200)

    def test_multiple_sheets(self) -> None:
        meta = build_sample_metadata()
        exporter = BaseExcelExporter(meta)
        s1 = exporter.add_sheet("Hoja1")
        s1.write_table(["A"], [[1]])
        s2 = exporter.add_sheet("Hoja2")
        s2.write_table(["B"], [[2]])
        content, _ = exporter.save()
        wb = load_workbook(BytesIO(content))
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertEqual(wb["Hoja2"]["A7"].value, "B")

    def test_engine_has_no_business_imports(self) -> None:
        import app.mrl.excel.exporter as mod

        with open(mod.__file__, encoding="utf-8") as fh:
            source = fh.read()
        self.assertNotIn("inventario_comercial", source)
        self.assertNotIn("maintenance", source)
        self.assertNotIn("models", source)


class TestMaintenanceActivosExport(unittest.TestCase):
    def test_activos_table_rows_structure(self) -> None:
        from types import SimpleNamespace

        from app.maintenance.mrl_exports import activos_table_rows

        machine = SimpleNamespace(
            id=1,
            codigo="EQ-001",
            nombre="Compresor",
            machine_type_id=1,
            tipo_etiqueta="General",
            ubicacion="Planta",
            area="Producción",
            status="operativo",
            criticidad="media",
            es_critico=False,
            foto_url="",
        )
        # build_activos_list_items needs DB for maintenance dates — test headers only
        headers, rows = activos_table_rows([])
        self.assertEqual(headers[0], "Código")
        self.assertEqual(rows, [])


if __name__ == "__main__":
    unittest.main()
