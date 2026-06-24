"""Exportaciones del módulo inventario comercial."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.currency import empresa_multimoneda, monedas_activas_de
from app.inventario_comercial.service import query_productos_bajo_stock
from app.models import Empresa


def _escribir_encabezado_empresa(ws, empresa: Empresa | None, hoy: date, num_cols: int) -> None:
    """Filas 1-6 con datos de la empresa; la fila 7 queda para encabezados de columnas."""
    nombre = ((empresa.razon_social if empresa else "") or "Empresa").strip()
    nit = (empresa.nit or "").strip() if empresa else ""
    partes_dir = []
    contacto = ""
    if empresa:
        partes_dir = [
            p.strip()
            for p in (empresa.direccion, empresa.ciudad, empresa.pais)
            if (p or "").strip()
        ]
        contacto_partes = []
        if (empresa.telefono or "").strip():
            contacto_partes.append(f"Tel: {empresa.telefono.strip()}")
        if (empresa.email or "").strip():
            contacto_partes.append(f"Email: {empresa.email.strip()}")
        contacto = " | ".join(contacto_partes)

    filas = [
        (nombre, Font(bold=True, size=14)),
        (f"NIT: {nit}" if nit else "", None),
        (f"Dirección: {', '.join(partes_dir)}" if partes_dir else "", None),
        (contacto, None),
        (f"Productos con bajo stock — {hoy:%d/%m/%Y}", Font(bold=True)),
        ("", None),
    ]
    for idx, (texto, fuente) in enumerate(filas, start=1):
        if not texto:
            continue
        cell = ws.cell(row=idx, column=1, value=texto)
        if fuente:
            cell.font = fuente
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=num_cols)


def excel_productos_bajo_stock(empresa_id: int) -> tuple[bytes, str]:
    """Genera un .xlsx con productos activos en bajo stock. Devuelve (contenido, nombre_archivo)."""
    empresa = Empresa.query.get(empresa_id)
    moneda_ref = (empresa.moneda if empresa else "COP") or "COP"
    multimoneda = empresa_multimoneda(empresa)
    monedas = monedas_activas_de(empresa) if multimoneda else [moneda_ref]
    productos = query_productos_bajo_stock(empresa_id).all()
    hoy = date.today()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bajo stock"

    headers = [
        "Referencia",
        "Nombre",
        "Marca",
        "Categoría",
        "Unidad",
        "Stock actual",
        "Stock mínimo",
        "Faltante",
        f"Precio compra ({moneda_ref})",
    ]
    if multimoneda:
        headers.extend(f"Precio venta ({m})" for m in monedas)
    else:
        headers.append(f"Precio venta ({moneda_ref})")

    _escribir_encabezado_empresa(ws, empresa, hoy, len(headers))

    header_row = 7
    for col_idx, titulo in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=titulo)
        cell.font = Font(bold=True)

    for p in productos:
        stock = int(p.stock or 0)
        minimo = int(p.stock_minimo or 0)
        faltante = max(0, minimo - stock)
        row = [
            p.sku,
            p.nombre,
            p.marca or "",
            p.categoria or "",
            p.unidad or "pza",
            stock,
            minimo,
            faltante,
            float(p.precio_compra or 0),
        ]
        if multimoneda:
            precios = p.precios_venta(moneda_ref)
            row.extend(float(precios.get(m, 0)) for m in monedas)
        else:
            row.append(float(p.precio_venta or 0))
        ws.append(row)

    for col_idx, _ in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)

    buf = BytesIO()
    wb.save(buf)
    nombre = f"productos_bajo_stock_{hoy:%Y%m%d}.xlsx"
    return buf.getvalue(), nombre
