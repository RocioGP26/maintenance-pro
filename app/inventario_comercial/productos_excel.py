"""Exportación e importación masiva de productos (.xlsx)."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app import db
from app.currency import empresa_multimoneda, monedas_activas_de
from app.inventario_comercial.service import _parse_float, _parse_int, guardar_producto_comercial
from app.models import Empresa, InvProducto

MAX_FILAS_IMPORT = 5000
MAX_BYTES_IMPORT = 10 * 1024 * 1024

_COL_REFERENCIA = "referencia"
_COL_NOMBRE = "nombre"


def _normalizar_texto(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _normalizar_clave(texto: str) -> str:
    raw = unicodedata.normalize("NFKD", (texto or "").strip().lower())
    sin_acentos = "".join(c for c in raw if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "_", sin_acentos).strip("_")


def _parse_activo(val: Any, default: bool = True) -> bool:
    raw = _normalizar_texto(val).lower()
    if not raw:
        return default
    if raw in ("0", "no", "n", "false", "falso", "inactivo", "inactiva"):
        return False
    if raw in ("1", "si", "sí", "s", "true", "verdadero", "activo", "activa", "yes", "y"):
        return True
    return default


def headers_producto_excel(empresa: Empresa | None) -> list[str]:
    moneda_ref = (empresa.moneda if empresa else "COP") or "COP"
    multimoneda = empresa_multimoneda(empresa)
    monedas = monedas_activas_de(empresa) if multimoneda else [moneda_ref]

    headers = [
        "Referencia",
        "Nombre",
        "Marca",
        "Categoría",
        "Subcategoría",
        "Unidad",
        "Stock",
        "Stock mínimo",
        f"Precio compra ({moneda_ref})",
    ]
    if multimoneda:
        headers.extend(f"Precio venta ({m})" for m in monedas)
    else:
        headers.append(f"Precio venta ({moneda_ref})")
    headers.append("Activo")
    return headers


def fila_producto_excel(producto: InvProducto, empresa: Empresa | None) -> list:
    moneda_ref = (empresa.moneda if empresa else "COP") or "COP"
    multimoneda = empresa_multimoneda(empresa)
    monedas = monedas_activas_de(empresa) if multimoneda else [moneda_ref]

    row: list = [
        producto.sku,
        producto.nombre,
        producto.marca or "",
        producto.categoria or "",
        producto.subcategoria or "",
        producto.unidad or "pza",
        int(producto.stock or 0),
        int(producto.stock_minimo or 0),
        float(producto.precio_compra or 0),
    ]
    if multimoneda:
        precios = producto.precios_venta(moneda_ref)
        row.extend(float(precios.get(m, 0)) for m in monedas)
    else:
        row.append(float(producto.precio_venta or 0))
    row.append("Sí" if producto.activo else "No")
    return row


def _ajustar_columnas(ws, num_cols: int) -> None:
    for col_idx in range(1, num_cols + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)


def _workbook_a_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_productos_catalogo(
    empresa_id: int,
    productos: list[InvProducto],
) -> tuple[bytes, str]:
    empresa = Empresa.query.get(empresa_id)
    headers = headers_producto_excel(empresa)
    hoy = date.today()

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    for col_idx, titulo in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = Font(bold=True)

    for producto in productos:
        ws.append(fila_producto_excel(producto, empresa))

    _ajustar_columnas(ws, len(headers))
    nombre = f"productos_catalogo_{hoy:%Y%m%d}.xlsx"
    return _workbook_a_bytes(wb), nombre


def excel_productos_plantilla(empresa_id: int) -> tuple[bytes, str]:
    empresa = Empresa.query.get(empresa_id)
    headers = headers_producto_excel(empresa)
    moneda_ref = (empresa.moneda if empresa else "COP") or "COP"
    multimoneda = empresa_multimoneda(empresa)
    monedas = monedas_activas_de(empresa) if multimoneda else [moneda_ref]

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    for col_idx, titulo in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = Font(bold=True)

    ejemplo = [
        "SKU-001",
        "Aceite motor 1L",
        "Castrol",
        "Lubricantes",
        "Sintéticos",
        "pza",
        10,
        5,
        8500.0,
    ]
    if multimoneda:
        ejemplo.extend(12000.0 for _ in monedas)
    else:
        ejemplo.append(12000.0)
    ejemplo.append("Sí")
    ws.append(ejemplo)

    inst = wb.create_sheet("Instrucciones")
    inst["A1"] = "Cómo importar productos"
    inst["A1"].font = Font(bold=True, size=12)
    lineas = [
        "1. Complete la hoja «Productos» (puede borrar la fila de ejemplo).",
        "2. Referencia y Nombre son obligatorios.",
        "3. Si la Referencia ya existe, el producto se actualiza; si no, se crea.",
        "4. Activo: use Sí o No.",
        f"5. Precios en {moneda_ref}" + (" y monedas adicionales según columnas." if multimoneda else "."),
    ]
    for idx, texto in enumerate(lineas, start=3):
        inst.cell(row=idx, column=1, value=texto)

    _ajustar_columnas(ws, len(headers))
    inst.column_dimensions["A"].width = 72
    return _workbook_a_bytes(wb), "plantilla_productos.xlsx"


def _mapear_columnas(headers: list[str], empresa: Empresa | None) -> dict[str, int]:
    moneda_ref = (empresa.moneda if empresa else "COP") or "COP"
    multimoneda = empresa_multimoneda(empresa)
    monedas = monedas_activas_de(empresa) if multimoneda else [moneda_ref]

    alias = {
        _COL_REFERENCIA: {_COL_REFERENCIA, "sku", "codigo", "codigo_referencia", "ref"},
        _COL_NOMBRE: {_COL_NOMBRE, "producto", "descripcion"},
        "marca": {"marca"},
        "categoria": {"categoria", "categoria_producto"},
        "subcategoria": {"subcategoria"},
        "unidad": {"unidad", "um"},
        "stock": {"stock", "stock_actual", "cantidad"},
        "stock_minimo": {"stock_minimo", "minimo", "stock_min"},
        "precio_compra": {"precio_compra", f"precio_compra_{moneda_ref.lower()}"},
        "activo": {"activo", "estado"},
    }
    if multimoneda:
        for m in monedas:
            alias[f"precio_venta_{m}"] = {
                f"precio_venta_{m.lower()}",
                f"precio_venta_{m}",
                f"venta_{m.lower()}",
            }
    else:
        alias["precio_venta"] = {
            "precio_venta",
            f"precio_venta_{moneda_ref.lower()}",
            "precio_venta",
        }

    clave_a_campo: dict[str, str] = {}
    for campo, keys in alias.items():
        for k in keys:
            clave_a_campo[k] = campo

    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        clave = _normalizar_clave(header)
        if not clave:
            continue
        if clave.startswith("precio_compra"):
            mapping.setdefault("precio_compra", idx)
            continue
        if clave.startswith("precio_venta"):
            moneda_match = re.search(r"\(([A-Z]{3})\)", header or "", re.I)
            if moneda_match:
                mapping[f"precio_venta_{moneda_match.group(1).upper()}"] = idx
            elif "precio_venta" in clave_a_campo.values() or clave in clave_a_campo:
                mapping.setdefault("precio_venta", idx)
            continue
        campo = clave_a_campo.get(clave)
        if campo:
            mapping.setdefault(campo, idx)
    return mapping


def _detectar_fila_encabezados(ws, empresa: Empresa | None) -> tuple[int, dict[str, int]] | None:
    for row_idx in range(1, min(ws.max_row, 25) + 1):
        headers = [
            _normalizar_texto(ws.cell(row=row_idx, column=c).value)
            for c in range(1, 31)
        ]
        while headers and not headers[-1]:
            headers.pop()
        if not any(headers):
            continue
        claves = {_normalizar_clave(v) for v in headers if v}
        if _COL_REFERENCIA in claves or "sku" in claves or "codigo" in claves:
            mapping = _mapear_columnas(headers, empresa)
            if _COL_REFERENCIA in mapping and _COL_NOMBRE in mapping:
                return row_idx, mapping
    return None


def _valor_celda(row: tuple, idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


@dataclass
class ResultadoImportProductos:
    creados: int = 0
    actualizados: int = 0
    omitidos: int = 0
    errores: list[str] = field(default_factory=list)


def importar_productos_desde_excel(
    empresa_id: int,
    contenido: bytes,
    *,
    nombre_archivo: str = "",
) -> ResultadoImportProductos:
    resultado = ResultadoImportProductos()
    if len(contenido) > MAX_BYTES_IMPORT:
        resultado.errores.append(f"El archivo supera el máximo de {MAX_BYTES_IMPORT // (1024 * 1024)} MB.")
        return resultado

    ext = (nombre_archivo or "").rsplit(".", 1)[-1].lower()
    if ext and ext not in ("xlsx", "xlsm"):
        resultado.errores.append("Solo se admiten archivos .xlsx.")
        return resultado

    empresa = Empresa.query.get(empresa_id)
    moneda_ref = (empresa.moneda if empresa else "COP") or "COP"
    multimoneda = empresa_multimoneda(empresa)
    monedas = monedas_activas_de(empresa) if multimoneda else [moneda_ref]

    try:
        wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    except Exception:
        resultado.errores.append("No se pudo leer el archivo Excel.")
        return resultado

    ws = wb["Productos"] if "Productos" in wb.sheetnames else wb.active
    detectado = _detectar_fila_encabezados(ws, empresa)
    if not detectado:
        wb.close()
        resultado.errores.append(
            "No se encontró la fila de encabezados (columnas Referencia y Nombre)."
        )
        return resultado

    header_row, mapping = detectado

    filas_datos = 0
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if filas_datos >= MAX_FILAS_IMPORT:
            resultado.errores.append(f"Se omitieron filas después del máximo de {MAX_FILAS_IMPORT}.")
            break

        sku = _normalizar_texto(_valor_celda(row, mapping.get(_COL_REFERENCIA)))
        nombre = _normalizar_texto(_valor_celda(row, mapping.get(_COL_NOMBRE)))
        if not sku and not nombre:
            continue
        if not sku or not nombre:
            resultado.errores.append(f"Fila {row_idx}: Referencia y Nombre son obligatorios.")
            resultado.omitidos += 1
            continue

        filas_datos += 1
        datos: dict[str, Any] = {
            "sku": sku,
            "nombre": nombre,
            "marca": _normalizar_texto(_valor_celda(row, mapping.get("marca"))),
            "categoria": _normalizar_texto(_valor_celda(row, mapping.get("categoria"))),
            "subcategoria": _normalizar_texto(_valor_celda(row, mapping.get("subcategoria"))),
            "unidad": _normalizar_texto(_valor_celda(row, mapping.get("unidad"))) or "pza",
            "stock": _parse_int(_valor_celda(row, mapping.get("stock"))),
            "stock_minimo": _parse_int(_valor_celda(row, mapping.get("stock_minimo"))),
            "precio_compra": _parse_float(_valor_celda(row, mapping.get("precio_compra"))),
            "activo": _parse_activo(_valor_celda(row, mapping.get("activo")), default=True),
        }
        if multimoneda:
            for m in monedas:
                datos[f"precio_venta_{m}"] = _parse_float(
                    _valor_celda(row, mapping.get(f"precio_venta_{m}"))
                )
        else:
            datos["precio_venta"] = _parse_float(
                _valor_celda(row, mapping.get("precio_venta"))
            )

        existente = InvProducto.query.filter_by(empresa_id=empresa_id, sku=sku).first()
        try:
            with db.session.begin_nested():
                if existente:
                    producto = guardar_producto_comercial(
                        empresa_id,
                        datos,
                        producto=existente,
                        permitir_stock_inicial=False,
                    )
                    if "stock" in mapping:
                        producto.stock = max(0, datos["stock"])
                    resultado.actualizados += 1
                else:
                    guardar_producto_comercial(
                        empresa_id,
                        datos,
                        permitir_stock_inicial=True,
                    )
                    resultado.creados += 1
                db.session.flush()
        except ValueError as exc:
            resultado.errores.append(f"Fila {row_idx}: {exc}")
            resultado.omitidos += 1
        except Exception:
            resultado.errores.append(f"Fila {row_idx}: no se pudo guardar el producto.")
            resultado.omitidos += 1

    wb.close()
    if filas_datos == 0 and not resultado.errores:
        resultado.errores.append("El archivo no contiene filas de productos.")
    return resultado
