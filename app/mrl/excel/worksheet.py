"""Helpers de hoja Excel MRL."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from app.mrl import colors
from app.mrl import constants
from app.mrl import typography
from app.mrl.excel.formatting import apply_institutional_meta_row
from app.mrl.metadata import MRLDocumentMeta
from app.mrl.utils.dates import format_datetime_latam


def write_institutional_header(
    ws: Worksheet,
    meta: MRLDocumentMeta,
    num_cols: int,
) -> int:
    """
    Escribe filas meta MRL (1–6) y retorna la fila del header de columnas (7).
    """
    fecha = format_datetime_latam(meta.generated_at, meta.timezone_name)
    lines: list[tuple[str, bool, int | None]] = [
        (meta.empresa_nombre, True, typography.EXCEL_FONT_SIZES.title),
        (f"NIT: {meta.empresa_nit}" if meta.empresa_nit else "", False, None),
        (meta.title, True, typography.EXCEL_FONT_SIZES.subtitle),
        (f"{meta.doc_code} · {meta.instance_code}", False, typography.EXCEL_FONT_SIZES.meta),
        (
            f"Módulo: {meta.module} · Usuario: {meta.usuario} · {fecha}",
            False,
            typography.EXCEL_FONT_SIZES.meta,
        ),
        (
            f"{constants.GENERATED_BY_LABEL} · MRL v{meta.mrl_version}",
            False,
            typography.EXCEL_FONT_SIZES.footer,
        ),
    ]
    row = constants.EXCEL_META_START_ROW
    for text, bold, size in lines:
        apply_institutional_meta_row(
            ws,
            row,
            text,
            num_cols=num_cols,
            bold=bold,
            size=size,
            color=colors.PRIMARY if bold and row == constants.EXCEL_META_START_ROW else colors.BODY,
        )
        row += 1
    return constants.EXCEL_HEADER_ROW


def freeze_below_header(ws: Worksheet) -> None:
    """Congela paneles debajo del header institucional."""
    ws.freeze_panes = ws.cell(row=constants.EXCEL_DATA_START_ROW, column=1)


def apply_auto_filter(ws: Worksheet, header_row: int, last_row: int, num_cols: int) -> None:
    """Activa filtros automáticos sobre la tabla."""
    if last_row < header_row or num_cols < 1:
        return
    from openpyxl.utils import get_column_letter

    start = f"A{header_row}"
    end = f"{get_column_letter(num_cols)}{last_row}"
    ws.auto_filter.ref = f"{start}:{end}"
