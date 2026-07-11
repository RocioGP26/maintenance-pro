"""Autoajuste de columnas Excel MRL."""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def auto_fit_columns(
    ws: Worksheet,
    num_cols: int,
    *,
    first_row: int = 1,
    last_row: int | None = None,
    min_width: float = 10.0,
    max_width: float = 40.0,
    padding: float = 2.0,
) -> None:
    """Ajusta ancho de columnas según contenido."""
    if num_cols < 1:
        return
    max_row = last_row or ws.max_row
    for col_idx in range(1, num_cols + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(first_row, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + padding, min_width), max_width)
