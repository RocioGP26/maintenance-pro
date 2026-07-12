"""Aplicación de estilos openpyxl · MRL-07 / MRL-05."""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.mrl import colors
from app.mrl import typography
from app.mrl.excel.utils import hex_to_openpyxl
from app.mrl.styles import MRLStyle, TableStyle


def _thin_border(color: str) -> Border:
    side = Side(style="thin", color=hex_to_openpyxl(color))
    return Border(left=side, right=side, top=side, bottom=side)


def font_excel(*, bold: bool = False, size: int | None = None, color: str | None = None) -> Font:
    return Font(
        name=typography.EXCEL_FONT_FAMILY,
        bold=bold,
        size=size or typography.EXCEL_FONT_SIZES.body,
        color=hex_to_openpyxl(color) if color else None,
    )


def fill_solid(color: str) -> PatternFill:
    return PatternFill(
        fill_type="solid",
        start_color=hex_to_openpyxl(color),
        end_color=hex_to_openpyxl(color),
    )


def apply_institutional_meta_row(
    ws: Worksheet,
    row: int,
    text: str,
    *,
    num_cols: int,
    bold: bool = False,
    size: int | None = None,
    color: str = colors.BODY,
) -> None:
    if not text:
        return
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font_excel(bold=bold, size=size, color=color)
    cell.alignment = Alignment(vertical="center")
    if num_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)


def apply_table_header_row(
    ws: Worksheet,
    row: int,
    headers: list[str],
    table_style: TableStyle | None = None,
) -> None:
    style = table_style or MRLStyle.table()
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = font_excel(
            bold=True,
            size=style.font_size,
            color=style.header_text,
        )
        cell.fill = fill_solid(style.header_background)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border(style.border_color)


def apply_body_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any,
    *,
    table_style: TableStyle | None = None,
    zebra: bool = False,
    number_format: str | None = None,
) -> None:
    style = table_style or MRLStyle.table()
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font_excel(size=style.font_size, color=style.body_text)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = _thin_border(style.border_color)
    if zebra and style.zebra_enabled:
        cell.fill = fill_solid(style.zebra_background)
    if number_format:
        cell.number_format = number_format


def apply_footer_row(
    ws: Worksheet,
    row: int,
    text: str,
    *,
    num_cols: int,
) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font_excel(size=typography.EXCEL_FONT_SIZES.footer, color=colors.MUTED)
    cell.alignment = Alignment(vertical="center")
    if num_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
